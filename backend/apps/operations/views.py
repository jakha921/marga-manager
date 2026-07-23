from datetime import timedelta
from decimal import ROUND_HALF_UP, Decimal

from django.db.models import Max, Sum
from django.db.models.functions import Coalesce
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import serializers as drf_serializers
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.core.mixins import TenantCreateMixin, TenantQuerySetMixin
from apps.core.permissions import IsKitchenUserOrAbove
from apps.kitchens.models import Kitchen
from apps.products.models import Product

from .filters import OperationEntryFilter
from .models import OperationEntry
from .serializers import OperationEntrySerializer

ZERO = Decimal("0")

BASIC_HISTORY_DAYS = 30


def _basic_history_cutoff(user):
    """Порог глубины истории для тарифа BASIC (None — без ограничений).

    Ограничение фичи тарифа: Basic видит последние 30 дней,
    Pro и SUPER_ADMIN — всю историю.
    """
    org = getattr(user, "organization", None)
    if user.role == "SUPER_ADMIN" or org is None or org.plan != "BASIC":
        return None
    return timezone.localdate() - timedelta(days=BASIC_HISTORY_DAYS)


def _clamp_date_from(user, date_from):
    """Не даёт date_from уйти глубже порога тарифа BASIC."""
    cutoff = _basic_history_cutoff(user)
    if cutoff is None or not date_from:
        return date_from
    return max(str(date_from), cutoff.isoformat())


class OperationViewSet(TenantQuerySetMixin, TenantCreateMixin, viewsets.ModelViewSet):
    """CRUD операций."""

    queryset = OperationEntry.objects.select_related("kitchen", "to_kitchen", "product").order_by(
        "-date", "-time", "-created_at", "-id"
    )
    serializer_class = OperationEntrySerializer
    permission_classes = [IsKitchenUserOrAbove]
    filterset_class = OperationEntryFilter
    search_fields = ["product__name"]

    def get_queryset(self):
        qs = super().get_queryset()
        cutoff = _basic_history_cutoff(self.request.user)
        if cutoff is not None:
            qs = qs.filter(date__gte=cutoff)
        return qs

    @action(detail=False, url_path="last-incoming/(?P<product_id>[^/.]+)")
    def last_incoming(self, request, product_id=None):
        """Последняя цена прихода для автозаполнения."""
        entry = (
            self.get_queryset()
            .filter(
                type=OperationEntry.Type.INCOMING,
                product_id=product_id,
                price__isnull=False,
            )
            .order_by("-date", "-time", "-created_at", "-id")
            .first()
        )

        if entry:
            unit_price = (
                (entry.price / entry.quantity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                if entry.quantity
                else None
            )
            return Response(
                {
                    "price": str(entry.price),
                    "unit_price": str(unit_price) if unit_price is not None else None,
                    "quantity": str(entry.quantity),
                    "unit": entry.unit,
                }
            )
        return Response({"price": None, "unit_price": None, "quantity": None, "unit": None})

    @action(detail=False, url_path="export", methods=["get"])
    def export_excel(self, request):
        """Экспорт операций в xlsx."""
        import io

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        qs = self.filter_queryset(self.get_queryset())[:10_000]

        wb = Workbook()
        ws = wb.active
        ws.title = "Operations"

        # Стили
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CBD5E1", end_color="CBD5E1", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        # Заголовки
        headers = ["Sana", "Mahsulot", "O'lchov", "Miqdor", "Narx za ed.", "Summa"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Данные
        total_sum = ZERO
        for row_idx, op in enumerate(qs, 2):
            price = op.price or ZERO
            qty = op.quantity or ZERO
            unit_price = price / qty if qty else ZERO
            total_sum += price

            values = [
                str(op.date),
                op.product.name if op.product else "—",
                op.unit,
                float(qty),
                float(unit_price),
                float(price),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                if col >= 4:
                    cell.number_format = "#,##0"

        # Итого
        last_row = qs.count() + 2
        ws.cell(row=last_row, column=5, value="JAMI").font = Font(bold=True)
        ws.cell(row=last_row, column=5).fill = total_fill
        ws.cell(row=last_row, column=5).border = thin_border
        total_cell = ws.cell(row=last_row, column=6, value=float(total_sum))
        total_cell.font = Font(bold=True)
        total_cell.fill = total_fill
        total_cell.border = thin_border
        total_cell.number_format = "#,##0"

        # Ширина колонок
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from django.http import HttpResponse

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=operations_export.xlsx"
        return response


def _get_tenant_qs(user):
    """Получить базовый queryset с фильтрацией по организации."""
    from rest_framework.exceptions import PermissionDenied

    qs = OperationEntry.objects.all()
    if user.role == "SUPER_ADMIN":
        return qs
    if not user.organization:
        raise PermissionDenied("Пользователь не привязан к организации.")
    return qs.filter(organization=user.organization)


class KitchenReportView(APIView):
    """GET /api/analytics/kitchen-report/ — финансовый отчёт по кухням."""

    permission_classes = [IsKitchenUserOrAbove]

    def get(self, request):
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        if not date_from or not date_to:
            raise drf_serializers.ValidationError(
                {"detail": "Параметры date_from и date_to обязательны."}
            )

        kitchen_filter = request.query_params.get("kitchen")
        user = request.user
        qs = _get_tenant_qs(user)
        # Тариф BASIC видит отчёты только за последние 30 дней
        date_from = _clamp_date_from(user, date_from)
        date_to = max(str(date_to), date_from) if date_to else date_to

        # Кухни для отчёта
        kitchens_qs = Kitchen.objects.all()
        if user.role != "SUPER_ADMIN" and user.organization:
            kitchens_qs = kitchens_qs.filter(organization=user.organization)
        if kitchen_filter:
            kitchens_qs = kitchens_qs.filter(id=kitchen_filter)

        kitchens_data = []
        totals = {
            "beginning_balance": ZERO,
            "incoming": ZERO,
            "end_balance": ZERO,
            "transfers_out": ZERO,
            "transfers_in": ZERO,
            "actual_expense": ZERO,
            "sales_revenue": ZERO,
            "markup_val": ZERO,
        }

        # 6 групповых запросов вместо 6N (по одному aggregate на кухню)
        def _agg_by_kitchen(queryset):
            return {
                row["kitchen"]: row["total"]
                for row in queryset.values("kitchen").annotate(total=Coalesce(Sum("price"), ZERO))
            }

        def _agg_by_to_kitchen(queryset):
            return {
                row["to_kitchen"]: row["total"]
                for row in queryset.values("to_kitchen").annotate(
                    total=Coalesce(Sum("price"), ZERO)
                )
            }

        def _balance_by_kitchen(on_or_before):
            """Остаток кухни: сумма последней DAILY-записи не позже указанной даты.

            Fallback нужен, потому что остаток фиксируется не каждый день —
            без него границы диапазона без записей дают нулевые остатки
            и бессмысленную отрицательную маржу.
            """
            latest_dates = (
                qs.filter(type=OperationEntry.Type.DAILY, date__lte=on_or_before)
                .values("kitchen")
                .annotate(last_date=Max("date"))
            )
            # ponytail: по запросу на кухню; лимит кухонь в тарифе <= 10 — норм
            return {
                row["kitchen"]: qs.filter(
                    type=OperationEntry.Type.DAILY,
                    kitchen=row["kitchen"],
                    date=row["last_date"],
                ).aggregate(total=Coalesce(Sum("price"), ZERO))["total"]
                for row in latest_dates
            }

        beginning_map = _balance_by_kitchen(date_from)
        incoming_map = _agg_by_kitchen(
            qs.filter(
                type=OperationEntry.Type.INCOMING,
                date__gte=date_from,
                date__lte=date_to,
            )
        )
        end_balance_map = _balance_by_kitchen(date_to)
        transfers_out_map = _agg_by_kitchen(
            qs.filter(
                type=OperationEntry.Type.TRANSFER,
                date__gte=date_from,
                date__lte=date_to,
            )
        )
        transfers_in_map = _agg_by_to_kitchen(
            qs.filter(
                type=OperationEntry.Type.TRANSFER,
                date__gte=date_from,
                date__lte=date_to,
                to_kitchen__isnull=False,
            )
        )
        sales_map = _agg_by_kitchen(
            qs.filter(
                type=OperationEntry.Type.SALE,
                date__gte=date_from,
                date__lte=date_to,
            )
        )

        for kitchen in kitchens_qs:
            kid = kitchen.id
            beginning_balance = beginning_map.get(kid, ZERO)
            incoming = incoming_map.get(kid, ZERO)
            end_balance = end_balance_map.get(kid, ZERO)
            transfers_out = transfers_out_map.get(kid, ZERO)
            transfers_in = transfers_in_map.get(kid, ZERO)
            sales_revenue = sales_map.get(kid, ZERO)
            actual_expense = (
                beginning_balance + incoming + transfers_in - transfers_out - end_balance
            )

            markup_val = sales_revenue - actual_expense
            markup_percent = (
                (markup_val / actual_expense * 100) if actual_expense > 0 else Decimal("0")
            )

            entry = {
                "kitchen_id": kid,
                "kitchen_name": kitchen.name,
                "beginning_balance": beginning_balance,
                "incoming": incoming,
                "end_balance": end_balance,
                "transfers_out": transfers_out,
                "transfers_in": transfers_in,
                "actual_expense": actual_expense,
                "sales_revenue": sales_revenue,
                "markup_val": markup_val,
                "markup_percent": round(markup_percent, 1),
            }
            kitchens_data.append(entry)

            # Аккумулируем итоги
            for key in totals:
                totals[key] += entry[key]

        totals["markup_percent"] = (
            round(totals["markup_val"] / totals["actual_expense"] * 100, 1)
            if totals["actual_expense"] > 0
            else Decimal("0")
        )

        # Проверяем, нужен ли xlsx export (используем "output" — "format" перехватывается DRF)
        fmt = request.query_params.get("output")
        if fmt == "xlsx":
            return self._export_xlsx(kitchens_data, totals, date_from, date_to)

        return Response({"kitchens": kitchens_data, "totals": totals})

    def _export_xlsx(self, kitchens_data, totals, date_from, date_to):
        """Экспорт kitchen report в xlsx."""
        import io

        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Kitchen Report"

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="CBD5E1", end_color="CBD5E1", fill_type="solid")
        total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Заголовок
        ws.merge_cells("A1:H1")
        title_cell = ws.cell(row=1, column=1, value="Asosiy Moliyaviy Hisobot")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=2, column=1, value=f"Davr: {date_from} — {date_to}").font = Font(italic=True)

        headers = [
            "Bo'lim",
            "Bosh. Qoldiq",
            "Kirim",
            "Xarajat",
            "Oxir. Qoldiq",
            "Sotuv",
            "Foyda",
            "%",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for i, k in enumerate(kitchens_data, 5):
            vals = [
                k["kitchen_name"],
                float(k["beginning_balance"]),
                float(k["incoming"]),
                float(k["actual_expense"]),
                float(k["end_balance"]),
                float(k["sales_revenue"]),
                float(k["markup_val"]),
                f"{k['markup_percent']}%",
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=col, value=v)
                cell.border = thin_border
                if isinstance(v, float):
                    cell.number_format = "#,##0"

        total_row = len(kitchens_data) + 5
        total_vals = [
            "JAMI (ITOGO)",
            float(totals["beginning_balance"]),
            float(totals["incoming"]),
            float(totals["actual_expense"]),
            float(totals["end_balance"]),
            float(totals["sales_revenue"]),
            float(totals["markup_val"]),
            f"{totals['markup_percent']}%",
        ]
        for col, v in enumerate(total_vals, 1):
            cell = ws.cell(row=total_row, column=col, value=v)
            cell.font = Font(bold=True)
            cell.fill = total_fill
            cell.border = thin_border
            if isinstance(v, float):
                cell.number_format = "#,##0"

        for letter, width in [
            ("A", 20),
            ("B", 15),
            ("C", 15),
            ("D", 15),
            ("E", 15),
            ("F", 15),
            ("G", 15),
            ("H", 10),
        ]:
            ws.column_dimensions[letter].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=kitchen_report.xlsx"
        return response


class OperationsSummaryView(APIView):
    """GET /api/analytics/operations-summary/ — итоги для QuickInput."""

    permission_classes = [IsKitchenUserOrAbove]

    def get(self, request):
        qs = _get_tenant_qs(request.user)
        cutoff = _basic_history_cutoff(request.user)
        if cutoff is not None:
            qs = qs.filter(date__gte=cutoff)

        # Фильтры
        op_type = request.query_params.get("type")
        kitchen = request.query_params.get("kitchen")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        search = request.query_params.get("search")

        if op_type:
            qs = qs.filter(type=op_type)
        if kitchen:
            qs = qs.filter(kitchen_id=kitchen)
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if search:
            qs = qs.filter(product__name__icontains=search)

        total_amount = qs.aggregate(total=Coalesce(Sum("price"), ZERO))["total"]
        count = qs.count()

        # Группировка quantity по unit
        unit_groups = qs.values("unit").annotate(total_qty=Sum("quantity")).order_by("unit")
        total_quantities = {g["unit"]: g["total_qty"] for g in unit_groups}

        return Response(
            {
                "total_amount": total_amount,
                "total_quantities": total_quantities,
                "count": count,
            }
        )


class SalesChartView(APIView):
    """GET /api/analytics/sales-chart/ — дневные продажи и закупки за период.

    Считается на сервере (не из клиентского кэша 1000 операций), поэтому
    не «недосчитывает» после роста числа операций. Возвращает плотный ряд
    по всем дням диапазона.
    """

    permission_classes = [IsKitchenUserOrAbove]

    def get(self, request):
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        if not date_from or not date_to:
            raise drf_serializers.ValidationError(
                {"detail": "Параметры date_from и date_to обязательны."}
            )

        user = request.user
        # Тариф BASIC видит только последние 30 дней
        date_from = _clamp_date_from(user, date_from)
        if date_to < date_from:
            date_to = date_from

        qs = _get_tenant_qs(user).filter(date__gte=date_from, date__lte=date_to)
        kitchen = request.query_params.get("kitchen")
        if kitchen and kitchen != "all":
            qs = qs.filter(kitchen_id=kitchen)

        by_date = {}
        for row in (
            qs.filter(type=OperationEntry.Type.SALE)
            .values("date")
            .annotate(total=Coalesce(Sum("price"), ZERO))
        ):
            by_date.setdefault(row["date"], {})["sales"] = row["total"]
        for row in (
            qs.filter(type=OperationEntry.Type.INCOMING)
            .values("date")
            .annotate(total=Coalesce(Sum("price"), ZERO))
        ):
            by_date.setdefault(row["date"], {})["purchases"] = row["total"]

        from datetime import date as date_cls

        start = date_cls.fromisoformat(str(date_from))
        end = date_cls.fromisoformat(str(date_to))
        series = []
        d = start
        while d <= end:
            entry = by_date.get(d, {})
            series.append(
                {
                    "date": d.isoformat(),
                    "sales": entry.get("sales", ZERO),
                    "purchases": entry.get("purchases", ZERO),
                }
            )
            d += timedelta(days=1)

        return Response({"series": series})


class ProductConsumptionView(APIView):
    """GET /api/analytics/product-consumption/<product_id>/ — расход продукта.

    Интервальный расчёт: между двумя записями остатка расход =
    прошлый остаток + приходы интервала (± трансферы для конкретной кухни)
    − текущий остаток. Значение приписывается дню записи остатка.
    Считается на сервере целиком, без ограничения кэша.
    """

    permission_classes = [IsKitchenUserOrAbove]

    def get(self, request, product_id):
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        if not date_from or not date_to:
            raise drf_serializers.ValidationError(
                {"detail": "Параметры date_from и date_to обязательны."}
            )

        user = request.user
        if user.role != "SUPER_ADMIN":
            if not user.organization:
                from rest_framework.exceptions import PermissionDenied

                raise PermissionDenied("Пользователь не привязан к организации.")
            get_object_or_404(Product, pk=product_id, organization=user.organization)
        else:
            get_object_or_404(Product, pk=product_id)

        date_from = _clamp_date_from(user, date_from)
        if date_to < date_from:
            date_to = date_from

        kitchen = request.query_params.get("kitchen")
        specific_kitchen = bool(kitchen and kitchen != "all")

        qs = _get_tenant_qs(user).filter(product_id=product_id)
        base = qs.filter(kitchen_id=kitchen) if specific_kitchen else qs

        # Остатки по дням (все известные, чтобы найти базу интервала до date_from)
        balance_by_date = {}
        for row in (
            base.filter(type=OperationEntry.Type.DAILY)
            .values("date")
            .annotate(total=Coalesce(Sum("quantity"), ZERO))
        ):
            balance_by_date[row["date"]] = row["total"]
        balance_dates = sorted(balance_by_date.keys())

        def sum_interval(op_type, after_excl, to_incl, by_to_kitchen=False):
            f = qs.filter(type=op_type, date__gt=after_excl, date__lte=to_incl)
            if specific_kitchen:
                f = (
                    f.filter(to_kitchen_id=kitchen)
                    if by_to_kitchen
                    else f.filter(kitchen_id=kitchen)
                )
            return f.aggregate(total=Coalesce(Sum("quantity"), ZERO))["total"]

        from datetime import date as date_cls

        start = date_cls.fromisoformat(str(date_from))
        end = date_cls.fromisoformat(str(date_to))
        series = []
        d = start
        while d <= end:
            value = ZERO
            if d in balance_by_date:
                prev = next((bd for bd in reversed(balance_dates) if bd < d), None)
                if prev is not None:
                    incoming = sum_interval(OperationEntry.Type.INCOMING, prev, d)
                    value = balance_by_date[prev] + incoming - balance_by_date[d]
                    if specific_kitchen:
                        t_out = sum_interval(OperationEntry.Type.TRANSFER, prev, d)
                        t_in = sum_interval(
                            OperationEntry.Type.TRANSFER, prev, d, by_to_kitchen=True
                        )
                        value = value + t_in - t_out
            series.append({"date": d.isoformat(), "value": value})
            d += timedelta(days=1)

        return Response({"series": series})
